# -*- coding: utf-8 -*-
"""
Created on Tue Jun 18 10:53:57 2024

@author: bsms9zh8
"""

import psychopy
import os.path as op
from psychopy import event, visual, gui, core
from intart_functions import sendParallelTrigger, readParallelTrigger
from create_stimuli_final import create_experimental_stimuli
from statistics import mean
import random
from psychopy.hardware import keyboard
import openpyxl
import keyboard as k 


### PARTICIPANT DATA - COMMENT OUT WHEN TESTING ### 

participant_info = {'Subject ID': '', 'Age': '', 'Sex': ['M', 'F'],
                    'Order': '', 'Counterbalance group': '', 'Random seed': ''}

myDlg = gui.DlgFromDict(participant_info, title="Sam gaze experiment")

if not myDlg.OK:
    core.quit()
    
order = int(participant_info['Order'])
subject_id = str(participant_info['Subject ID'])
random_seed = int(participant_info['Random seed'])
age = str(participant_info['Age'])
sex = str(participant_info['Sex'])
counterbalance_group = int(participant_info['Counterbalance group'])

### VARIABLES, PATHS, FUNCTIONS ### 

pport_address_spike = 16360
trigger_code_spike = 255

experiment_clock = core.MonotonicClock() # the main clock used in the experiment 
countdown = core.CountdownTimer()

win = visual.Window(color = '#000000', fullscr = True, monitor="testMonitor", units="pix")
win.mouseVisible = False

fr = win.getActualFrameRate() # getting frame rate of the monitor 
one_frame = 1/fr
inc = 0.5 # variable controlling the rating scale 

# variables defining the number of ticks and their spacing in the confidence scale 
num_ticks = 11
tick_length = 20  # Length of each horizontal tick in pixels
spacing = 25  # Spacing between ticks

# function that creates some stimuli 

fixation_cross_horizontal, fixation_cross_vertical, training_instructions_one, training_instructions_two, training_instructions_three, training_instructions_four, training_instructions_five, training_instructions_six, training_instructions_seven, training_instructions_eight, training_instructions_nine, training_complete, block_complete, confidence_instructions, high_confidence, low_confidence, vertical_line, line_length, f1_0v_0h, f1_0v_5h, f1_0v_m5h, f1_0v_10h, f1_0v_m10h, f1_0v_15h, f1_0v_m15h, f1_10v_0h, f1_m10v_0h, f1_10v_5h, f1_10v_m5h, f1_m10v_5h, f1_m10v_m5h, f1_10v_10h, f1_10v_m10h, f1_m10v_10h, f1_m10v_m10h, f1_10v_15h, f1_10v_m15h, f1_m10v_15h, f1_m10v_m15h, f2_0v_0h, f2_0v_5h, f2_0v_m5h, f2_0v_10h, f2_0v_m10h, f2_0v_15h, f2_0v_m15h, f2_10v_0h, f2_m10v_0h, f2_10v_5h, f2_10v_m5h, f2_m10v_5h, f2_m10v_m5h, f2_10v_10h, f2_10v_m10h, f2_m10v_10h, f2_m10v_m10h, f2_10v_15h, f2_10v_m15h, f2_m10v_15h, f2_m10v_m15h, f3_0v_0h, f3_0v_5h, f3_0v_m5h, f3_0v_10h, f3_0v_m10h, f3_0v_15h, f3_0v_m15h, f3_10v_0h, f3_m10v_0h, f3_10v_5h, f3_10v_m5h, f3_m10v_5h, f3_m10v_m5h, f3_10v_10h, f3_10v_m10h, f3_m10v_10h, f3_m10v_m10h, f3_10v_15h, f3_10v_m15h, f3_m10v_15h, f3_m10v_m15h = create_experimental_stimuli(win)

### SETTING UP THE EXPERIMENT STRUCTURE ### 

# 21 different gaze directions per face
# In each 'rep' we'll present the 21 different gaze directions at both systole and diastole i.e., 42 trials 
# We will then repeat each 'rep' 3 times, so that a given gaze direction is presented 3 times at both systole and diastole

task_order_1 = [1,2,2,1,2,1,1,2,1,2,2,1,2,1,2,1,1,2,1,2,2,1,2,1,1,2,1,2,1,2,2,1,2,1,1,2,1,2,2,1,2,1] # 1 = systole, 2 = diastole 
task_order_2 = [2,1,1,2,1,2,2,1,2,1,1,2,1,2,1,2,2,1,2,1,1,2,1,2,2,1,2,1,2,1,1,2,1,2,2,1,2,1,1,2,1,2] 

rep_length = len(task_order_1) # the number of trials in a rep 
number_reps = 3
trials_per_block = len(task_order_1) * number_reps

order_1_structure = [task_order_1, task_order_2, task_order_1, task_order_2, task_order_1] # will counterbalance 
order_2_structure = [task_order_2, task_order_1, task_order_2, task_order_1, task_order_2]

# Strings for recording gaze directions 
str_0v_0h = '0v_0h'
str_0v_5h = '0v_5h'
str_0v_m5h = '0v_m5h'
str_0v_10h = '0v_10h'
str_0v_m10h = '0v_m10h'
str_0v_15h = '0v_15h'
str_0v_m15h = '0v_m15h'
str_10v_0h = '10v_0h'
str_m10v_0h = 'm10v_0h'
str_10v_5h = '10v_5h'
str_10v_m5h = '10v_m5h'
str_m10v_5h = 'm10v_5h'
str_m10v_m5h = 'm10v_m5h'
str_10v_10h = '10v_10h'
str_10v_m10h = '10v_m10h'
str_m10v_10h = 'm10v_10h'
str_m10v_m10h = 'm10v_m10h'
str_10v_15h = '10v_15h'
str_10v_m15h = '10v_m15h'
str_m10v_15h = 'm10v_15h'
str_m10v_m15h = 'm10v_m15h'

### FUNCTION FOR HEARTBEAT DETECTION ###

def listenforHR_basic(): 
    pport_address_in = 16361 # set to 889 for psychopys lab tests 16377        
    while True:
        signal_in = readParallelTrigger(pport_address_in)
        if signal_in > 63: 
            time_hb_prev = experiment_clock.getTime()
            break
    return time_hb_prev


### SAVE PATH ### 

data_path = 'D:\\PhD\\cardiac_gaze\\data'
subject_id_file_name = subject_id + '_data.xlsx'
save_file = op.join(data_path, subject_id_file_name)

data_wb = openpyxl.Workbook() # opening the excel file now 

faces_strings = ['face_1', 'face_2', 'face_3']
variable_strings = ['gaze_angle', 'systole_diastole', 'perceptual_response_x', 'perceptual_response_y', 'distance_from_centre', 'confidence_response', 'r_peak_time_1', 'r_peak_time_2', 'r_peak_time_3', 'r_peak_time_4', 'r_peak_time_5', 'mean_rr_int', 'predicted_rr_int']

for ws_name in range(len(faces_strings)):
    data_wb.create_sheet(faces_strings[ws_name]) # creating a sheet to save each block 
    
    data_sheet = data_wb[faces_strings[ws_name]]
    
    for header in range(len(variable_strings)): # adding the headers
        data_sheet.cell(1, header+1, variable_strings[header])


### INSTRUCTIONS ### 

training_instructions_one.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
training_instructions_two.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
training_instructions_three.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
training_instructions_four.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
training_instructions_five.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
training_instructions_six.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
training_instructions_seven.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
training_instructions_eight.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()
# training_instructions_nine.draw()
# win.flip()
# event.waitKeys(keyList=['space'])
# win.flip()

### INSERT TRAINING BLOCK HERE ### 

### EXPERIMENTAL BLOCK 1 ###

gaze_directions = [] # containing gaze directions per trial for a block

systole_diastole = [] # whether the trial is systole or diastole for a block
perceptual_responses_x = [] # perceptual responses per trial for a block, x-coordinate
perceptual_responses_y = [] # perceptual responses per trial for a block, y-coordinate 
distances_from_direct = [] # perceptual response, direct from centre 
confidence_responses = [] # confidence responses per trial for a block
r_peak_times_lists = [] # r-peak times per trial for a block
mean_rr_intervals_list = [] # rr-intervals per trial for a block 
predicted_rr_intervals_list = [] # Predicted RR-intervals per trial for a block 
 
countdown.reset()

fixation_cross_horizontal.draw()
fixation_cross_vertical.draw()

win.flip()

while True:
    if countdown.getTime() < (-2 + one_frame):
        break 
    
for rep in range(number_reps): # each gaze direction is shown five times per systole/diastole i.e., 5 reps per face     

    # Resetting the stimuli for this rep. Each rep consists of showing each gaze direction at systole and diastole
    face_1_stim = [f1_0v_0h, f1_0v_5h, f1_0v_m5h, f1_0v_10h, f1_0v_m10h, f1_0v_15h, f1_0v_m15h, f1_10v_0h, f1_m10v_0h, f1_10v_5h, f1_10v_m5h, f1_m10v_5h, f1_m10v_m5h, f1_10v_10h, f1_10v_m10h, f1_m10v_10h, f1_m10v_m10h, f1_10v_15h, f1_10v_m15h, f1_m10v_15h, f1_m10v_m15h]
    face_1_stim_d = [f1_0v_0h, f1_0v_5h, f1_0v_m5h, f1_0v_10h, f1_0v_m10h, f1_0v_15h, f1_0v_m15h, f1_10v_0h, f1_m10v_0h, f1_10v_5h, f1_10v_m5h, f1_m10v_5h, f1_m10v_m5h, f1_10v_10h, f1_10v_m10h, f1_m10v_10h, f1_m10v_m10h, f1_10v_15h, f1_10v_m15h, f1_m10v_15h, f1_m10v_m15h]
    face_1_stim_s = [f1_0v_0h, f1_0v_5h, f1_0v_m5h, f1_0v_10h, f1_0v_m10h, f1_0v_15h, f1_0v_m15h, f1_10v_0h, f1_m10v_0h, f1_10v_5h, f1_10v_m5h, f1_m10v_5h, f1_m10v_m5h, f1_10v_10h, f1_10v_m10h, f1_m10v_10h, f1_m10v_m10h, f1_10v_15h, f1_10v_m15h, f1_m10v_15h, f1_m10v_m15h]

    
    if order == 1: # the order in which this rep occurs 
        rep_order = order_1_structure[rep]
    elif order == 2:
        rep_order = order_2_structure[rep]
    
    # Entering the main stimulus presentation 
    for trial in range(rep_length):
        
        r_peak_counter = 0 # for tracking how many R-peaks have been detected this trial so far - remove when collecting actual data 
        
        r_peak_times = [] # the list that we'll use for each trial to measure R-peak times
        rr_intervals = [] 
        
        if ((rep_order == task_order_1) and (task_order_1[trial] == 1)) or ((rep_order == task_order_2) and task_order_2[trial] == 1): # a systole trial 
        
            face_this_trial = random.choice(face_1_stim_s) # choosing which stimulus will be presented this trial
            face_1_stim_s.remove(face_this_trial) # removing the chosen simulus, so that it won't be presented this rep again 
            face_this_trial.draw() # drawing the chosen stimulus 
            
            systole_diastole.append(1) # systole = 1, diastole = 2
            
            # Adding to the gaze direction list the face direction presented this trial
            
            if face_this_trial == f1_0v_0h:
                gaze_directions.append(str_0v_0h)
            elif face_this_trial == f1_0v_5h:
                gaze_directions.append(str_0v_5h)
            elif face_this_trial == f1_0v_m5h:
                gaze_directions.append(str_0v_m5h)
            elif face_this_trial == f1_0v_10h:
                gaze_directions.append(str_0v_10h)
            elif face_this_trial == f1_0v_m10h:
                gaze_directions.append(str_0v_m10h)
            elif face_this_trial == f1_0v_15h:
                gaze_directions.append(str_0v_15h)
            elif face_this_trial == f1_0v_m15h:
                gaze_directions.append(str_0v_m15h)
            elif face_this_trial == f1_10v_0h:
                gaze_directions.append(str_10v_0h)
            elif face_this_trial == f1_m10v_0h:
                gaze_directions.append(str_m10v_0h)
            elif face_this_trial == f1_10v_5h:
                gaze_directions.append(str_10v_5h)
            elif face_this_trial == f1_10v_m5h:
                gaze_directions.append(str_10v_m5h)
            elif face_this_trial == f1_m10v_5h:
                gaze_directions.append(str_m10v_5h)
            elif face_this_trial == f1_m10v_m5h:
                gaze_directions.append(str_m10v_m5h)
            elif face_this_trial == f1_10v_10h:
                gaze_directions.append(str_10v_10h)
            elif face_this_trial == f1_10v_m10h:
                gaze_directions.append(str_10v_m10h)
            elif face_this_trial == f1_m10v_10h:
                gaze_directions.append(str_m10v_10h)
            elif face_this_trial == f1_m10v_m10h:
                gaze_directions.append(str_m10v_m10h)
            elif face_this_trial == f1_10v_15h:
                gaze_directions.append(str_10v_15h)
            elif face_this_trial == f1_10v_m15h:
                gaze_directions.append(str_10v_m15h)
            elif face_this_trial == f1_m10v_15h:
                gaze_directions.append(str_m10v_15h)
            elif face_this_trial == f1_m10v_m15h:
                gaze_directions.append(str_m10v_m15h)
                
            ### R-PEAK DETECTION ### 
            
            # For diastole trials we're going to be gathering 3 RR-intervals, so waiting for 4 R-peaks, then presenting
            # the stimulus just before the 5th R-peak. 
            
            # We'll do the same for systole trials to keep the ISIs similar, i.e., detecting 5 R-peaks, and presented 250ms after the 5th  
            
            for peak in range(5): 
                r_peak_time = listenforHR_basic()
                r_peak_times.append(r_peak_time)
                countdown.reset()
                r_peak_counter += 1 
                
                # A potential problem has been that when the threshold is crossed (i.e., an R-peak), the listenforHR_basic doesn't only send once, but during the
                # whole time that the signal crosses the threshold. So I'm introducing a delay during the sensing period, in order to pause everything
                # Until the signal has gone back below the threshold 
                
                if r_peak_counter < 5:
                    core.wait(0.5)
            
            while True: # waiting 250ms after R-peak has been detected
                if countdown.getTime() < (-0.250 + one_frame):
                    break
            
            win.flip()  
            countdown.reset()
            sendParallelTrigger(pport_address_spike, trigger_code_spike)
        
            while True:
                if countdown.getTime() < (-0.04 + one_frame): #40ms presentation 
                    break
                
            win.flip() # remove stimulus
            
            ### FOR SYSTOLE TRIALS, CALCULATING THE MEAN RR-INTERVAL AND ADDING RR-INTERVALS TO LIST ### 
            
            for peak in range(len(r_peak_times) - 1): # going through the R-peak list
                rr_interval = r_peak_times[peak + 1] - r_peak_times[peak] # R-peak, index + 1, minus R-peak, index
                rr_intervals.append(rr_interval)
                
            mean_rr_int = mean(rr_intervals)
            
            
        elif ((rep_order == task_order_1) and (task_order_1[trial] == 2)) or ((rep_order == task_order_2) and (task_order_2[trial] == 2)): # a diastole trial 
        
            face_this_trial = random.choice(face_1_stim_d) # choosing which stimulus will be presented this trial
            face_1_stim_d.remove(face_this_trial) # removing the chosen simulus, so that it won't be presented this rep again 
            face_this_trial.draw()  # drawing the chosen stimulus 
            
            systole_diastole.append(2)
            
            if face_this_trial == f1_0v_0h:
                gaze_directions.append(str_0v_0h)
            elif face_this_trial == f1_0v_5h:
                gaze_directions.append(str_0v_5h)
            elif face_this_trial == f1_0v_m5h:
                gaze_directions.append(str_0v_m5h)
            elif face_this_trial == f1_0v_10h:
                gaze_directions.append(str_0v_10h)
            elif face_this_trial == f1_0v_m10h:
                gaze_directions.append(str_0v_m10h)
            elif face_this_trial == f1_0v_15h:
                gaze_directions.append(str_0v_15h)
            elif face_this_trial == f1_0v_m15h:
                gaze_directions.append(str_0v_m15h)
            elif face_this_trial == f1_10v_0h:
                gaze_directions.append(str_10v_0h)
            elif face_this_trial == f1_m10v_0h:
                gaze_directions.append(str_m10v_0h)
            elif face_this_trial == f1_10v_5h:
                gaze_directions.append(str_10v_5h)
            elif face_this_trial == f1_10v_m5h:
                gaze_directions.append(str_10v_m5h)
            elif face_this_trial == f1_m10v_5h:
                gaze_directions.append(str_m10v_5h)
            elif face_this_trial == f1_m10v_m5h:
                gaze_directions.append(str_m10v_m5h)
            elif face_this_trial == f1_10v_10h:
                gaze_directions.append(str_10v_10h)
            elif face_this_trial == f1_10v_m10h:
                gaze_directions.append(str_10v_m10h)
            elif face_this_trial == f1_m10v_10h:
                gaze_directions.append(str_m10v_10h)
            elif face_this_trial == f1_m10v_m10h:
                gaze_directions.append(str_m10v_m10h)
            elif face_this_trial == f1_10v_15h:
                gaze_directions.append(str_10v_15h)
            elif face_this_trial == f1_10v_m15h:
                gaze_directions.append(str_10v_m15h)
            elif face_this_trial == f1_m10v_15h:
                gaze_directions.append(str_m10v_15h)
            elif face_this_trial == f1_m10v_m15h:
                gaze_directions.append(str_m10v_m15h)
            
            ### R-PEAK DETECTION AND RR-INTERVAL CALCULATION ### 
            
            for peak in range(4): # gathering 4 R-peaks
                r_peak_time = listenforHR_basic()
                r_peak_times.append(r_peak_time)
                countdown.reset()
                r_peak_counter += 1
            
                
                if peak < r_peak_counter < 4: # again stopping multiple signals being sent per R-peak 
                    core.wait(0.5)

            ### CALCULATING RR-INTERVALS ### 
            
            for peak in range(len(r_peak_times) - 1): # going through the R-peak list
                rr_interval = r_peak_times[peak + 1] - r_peak_times[peak] # R-peak, index + 1, minus R-peak, index
                rr_intervals.append(rr_interval)
                
            mean_rr_int = mean(rr_intervals) # this is useful to know to check for faulty trials 
            
            predicted_rr_int = (rr_intervals[-1]*0.5) + (rr_intervals[-2]*0.3) + (rr_intervals[-3]*0.2)
            # ^ I'm calculating the predicted RR-interval by weighting the calculated RR-intervals, most recent with the highest weighting 
            
            ### DOING SOME BASIC ADJUSTMENTS BASED OFF THE MEAN RR-INTERVALS ### 
            
            if predicted_rr_int >= 1: # i.e., less than 60bpm
            
                while True:
                    if countdown.getTime() < (-predicted_rr_int + 0.150 + one_frame): # we present the stimulus 150ms before the next predicted R-peak
                        break
            
            elif (1 > predicted_rr_int) and (predicted_rr_int >= 0.86): #i.e., between 60 and 70bpm
                
                while True:
                    if countdown.getTime() < (-predicted_rr_int + 0.135 + one_frame): # present the stimulus 0.135ms before the next predicted R-peak
                        break
                    
            elif predicted_rr_int < 0.86: # i.e., greater than 70bpm
            
                while True: 
                    if countdown.getTime() < (-predicted_rr_int + 0.125 + one_frame): # present the stimulus 0.125ms before the next predicted R-peak 
                        break                
            
            win.flip() # Presenting the face 
            countdown.reset() # resetting the countdown timer now
            sendParallelTrigger(pport_address_spike, trigger_code_spike)
            
            while True:
                if countdown.getTime() < (-0.04 + one_frame):
                    break
                
            win.flip()
            
        ### GETTING GAZE DIRECTION RESPONSE ### 
        
        countdown.reset()
        
        circle = psychopy.visual.Circle(win, fillColor = "white", radius = 150)
        cross = psychopy.visual.shape.ShapeStim(win, fillColor = "red", vertices = 'cross', size = 25)
        
        # Function to check if the cross is within the circle
        def is_within_circle(cross_pos, circle_radius):
            return (cross_pos[0]**2 + cross_pos[1]**2) <= circle_radius**2
        
        cross_pos = [0,0]
        cross.setPos(cross_pos)
        
        while True:
            if countdown.getTime() < (-0.150 + one_frame): # A blank screen for 150ms after the presentation of the face 
                break
            
        cross.draw()
        circle.draw()
        win.flip()
        
     
        while True:
        
            if k.is_pressed("left"):
                new_pos = [cross_pos[0] - 3, cross_pos[1]]
            elif k.is_pressed("right"):
                new_pos = [cross_pos[0] + 3, cross_pos[1]]
            elif k.is_pressed("up"):
                new_pos = [cross_pos[0], cross_pos[1] + 3]
            elif k.is_pressed("down"):
                new_pos = [cross_pos[0], cross_pos[1] - 3]
            elif k.is_pressed("space"):
                break
            else:
                new_pos = cross_pos
            
            # Check if the new position is within the circle
            if is_within_circle(new_pos, 150):
                cross_pos = new_pos
                cross.setPos(cross_pos)
        
            # Draw the circle and the cross
            circle.draw()
            cross.draw()
            win.flip()
           
        response_coords_x = cross_pos[0]
        response_coords_y = cross_pos[1]
        
        distance_from_direct = (response_coords_x**2 + response_coords_y**2)**0.5 
        
        ### GETTING CONFIDENCE RESPONSE ### 
        
        countdown.reset()
      
        while True:
            if countdown.getTime() < (-0.25 + one_frame): # A blank screen for 250ms after the first perceptual rating 
                break
            
        # Followed by the confidence rating scale 
        
        triangle_position = 0
        space_pressed = False
        while space_pressed == False:
            
            # Draw the vertical line and confidence instructions
            vertical_line.draw()
            confidence_instructions.draw()
            high_confidence.draw()
            low_confidence.draw()
            triangle = visual.Polygon(win, edges = 3, radius = 15, pos = [-50, triangle_position])
            triangle.draw()

            
            for i in range(num_ticks):
                # Calculate the y position of each tick
                y_pos = -line_length / 2 + i * spacing
                # Create the horizontal tick
                tick = visual.Line(
                    win,
                    start=(-tick_length / 2, y_pos),
                    end=(tick_length / 2, y_pos),
                    lineColor='white')
                # Draw the tick
                tick.draw()
                
            win.flip()
            
            keys = event.waitKeys(keyList = ['up', 'down', 'space'])
            
            if (keys[0] == 'up') and (triangle_position != 125):
                triangle_position += 25
                
            elif (keys[0] == 'up') and (triangle_position == 125):
                triangle_position = 125
                
            elif (keys[0] == 'down') and (triangle_position != -125):
                triangle_position -= 25
            
            elif (keys[0] == 'down') and (triangle_position == -125):
                triangle_position == -25
                
            elif keys[0] == 'space':
                break 
            
        if triangle_position == 0:
            confidenceRating = 5
        elif triangle_position == 25:
            confidenceRating = 6
        elif triangle_position == 50:
            confidenceRating = 7
        elif triangle_position == 75:
            confidenceRating = 8
        elif triangle_position == 100:
            confidenceRating = 9
        elif triangle_position == 125:
            confidenceRating = 10
        elif triangle_position == -25:
            confidenceRating = 4
        elif triangle_position == -50:
            confidenceRating = 3
        elif triangle_position == -75:
            confidenceRating = 2
        elif triangle_position == -100:
            confidenceRating = 1
        elif triangle_position == -125:
            confidenceRating = 0 
                
        # Drawing the fixation cross for the next trial, will be shown for the next R-peak sensing period     
        
        win.flip()
        
        fixation_cross_horizontal.draw()
        fixation_cross_vertical.draw()
        
        # Recording the info for this trial 
        
        perceptual_responses_x.append(response_coords_x)
        perceptual_responses_y.append(response_coords_y)
        distances_from_direct.append(distance_from_direct)
        confidence_responses.append(confidenceRating)
        r_peak_times_lists.append(r_peak_times)
        mean_rr_intervals_list.append(mean_rr_int)
        
        if systole_diastole[-1] == 1: # a systole trial 
            predicted_rr_intervals_list.append(0) # just to fill the list (during systole trials we don't calculate a predicted rr int)
        
        elif systole_diastole[-1] == 2: # a diastole trial 
            predicted_rr_intervals_list.append(predicted_rr_int)
            
        win.flip()

### SAVING DATA FOR FACE 1 ### 

data_sheet = data_wb['face_1'] 

# Adding data to worksheet 

for trial in range(trials_per_block):
    # Adding gaze direction data for a trial 
    data_sheet.cell(trial+2, 1, gaze_directions[trial])
    # Adding systole/diastole data for a trial 
    data_sheet.cell(trial+2, 2, systole_diastole[trial])
    # Adding perceptual rating, x-coordinate
    data_sheet.cell(trial+2, 3, perceptual_responses_x[trial])
    # Adding perceptual rating, y-coordinate
    data_sheet.cell(trial+2, 4, perceptual_responses_y[trial])
    # Adding perceptual rating, distance from centre 
    data_sheet.cell(trial+2, 5, distances_from_direct[trial])
    # Adding confidence rating 
    data_sheet.cell(trial+2, 6, confidence_responses[trial])
    
    if systole_diastole[trial] == 1: # 5 r-peaks detected during a systole trial 
        
        data_sheet.cell(trial+2, 7, r_peak_times_lists[trial][0]) # first R-peak
        data_sheet.cell(trial+2, 8, r_peak_times_lists[trial][1]) # second R-peak...
        data_sheet.cell(trial+2, 9, r_peak_times_lists[trial][2]) 
        data_sheet.cell(trial+2, 10, r_peak_times_lists[trial][3]) 
        data_sheet.cell(trial+2, 11, r_peak_times_lists[trial][4]) 
    
    elif systole_diastole[trial] == 2: # 4 r-peaks detected during a diastole trial 
        
        data_sheet.cell(trial+2, 7, r_peak_times_lists[trial][0]) # first R-peak
        data_sheet.cell(trial+2, 8, r_peak_times_lists[trial][1]) # second R-peak...
        data_sheet.cell(trial+2, 9, r_peak_times_lists[trial][2]) 
        data_sheet.cell(trial+2, 10, r_peak_times_lists[trial][3]) 
        data_sheet.cell(trial+2, 13, predicted_rr_intervals_list[trial]) # predicted RR-interval added

    data_sheet.cell(trial+2, 12, mean_rr_intervals_list[trial]) # mean rr_interval added 
    
    
### EXPERIMENTAL BLOCK 2 ###

#Showing block complete dialogue

block_complete.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()

gaze_directions = [] # containing gaze directions per trial for a block

systole_diastole = [] # whether the trial is systole or diastole for a block
perceptual_responses_x = [] # perceptual responses per trial for a block, x-coordinate
perceptual_responses_y = [] # perceptual responses per trial for a block, y-coordinate 
distances_from_direct = [] # perceptual response, direct from centre 
confidence_responses = [] # confidence responses per trial for a block
r_peak_times_lists = [] # r-peak times per trial for a block
mean_rr_intervals_list = [] # rr-intervals per trial for a block 
predicted_rr_intervals_list = [] # Predicted RR-intervals per trial for a block 
 
countdown.reset()

fixation_cross_horizontal.draw()
fixation_cross_vertical.draw()

win.flip()

while True:
    if countdown.getTime() < (-2 + one_frame):
        break 

for rep in range(number_reps): # each gaze direction is shown five times per systole/diastole i.e., 5 reps per face     

    # Resetting the stimuli for this rep. Each rep consists of showing each gaze direction at systole and diastole
    face_2_stim = [f2_0v_0h, f2_0v_5h, f2_0v_m5h, f2_0v_10h, f2_0v_m10h, f2_0v_15h, f2_0v_m15h, f2_10v_0h, f2_m10v_0h, f2_10v_5h, f2_10v_m5h, f2_m10v_5h, f2_m10v_m5h, f2_10v_10h, f2_10v_m10h, f2_m10v_10h, f2_m10v_m10h, f2_10v_15h, f2_10v_m15h, f2_m10v_15h, f2_m10v_m15h]
    face_2_stim_d = [f2_0v_0h, f2_0v_5h, f2_0v_m5h, f2_0v_10h, f2_0v_m10h, f2_0v_15h, f2_0v_m15h, f2_10v_0h, f2_m10v_0h, f2_10v_5h, f2_10v_m5h, f2_m10v_5h, f2_m10v_m5h, f2_10v_10h, f2_10v_m10h, f2_m10v_10h, f2_m10v_m10h, f2_10v_15h, f2_10v_m15h, f2_m10v_15h, f2_m10v_m15h]
    face_2_stim_s = [f2_0v_0h, f2_0v_5h, f2_0v_m5h, f2_0v_10h, f2_0v_m10h, f2_0v_15h, f2_0v_m15h, f2_10v_0h, f2_m10v_0h, f2_10v_5h, f2_10v_m5h, f2_m10v_5h, f2_m10v_m5h, f2_10v_10h, f2_10v_m10h, f2_m10v_10h, f2_m10v_m10h, f2_10v_15h, f2_10v_m15h, f2_m10v_15h, f2_m10v_m15h]
    
    if order == 1: # the order in which this rep occurs 
        rep_order = order_1_structure[rep]
    elif order == 2:
        rep_order = order_2_structure[rep]
    
    # Entering the main stimulus presentation 
    for trial in range(rep_length):
        
        r_peak_counter = 0 # for tracking how many R-peaks have been detected this trial so far - remove when collecting actual data 
        
        r_peak_times = [] # the list that we'll use for each trial to measure R-peak times
        rr_intervals = [] 
        
        if ((rep_order == task_order_1) and (task_order_1[trial] == 1)) or ((rep_order == task_order_2) and task_order_2[trial] == 1): # a systole trial 
        
            face_this_trial = random.choice(face_2_stim_s) # choosing which stimulus will be presented this trial
            face_2_stim_s.remove(face_this_trial) # removing the chosen simulus, so that it won't be presented this rep again 
            face_this_trial.draw() # drawing the chosen stimulus 
            
            systole_diastole.append(1) # systole = 1, diastole = 2
            
            # Adding to the gaze direction list the face direction presented this trial
            
            if face_this_trial == f2_0v_0h:
                gaze_directions.append(str_0v_0h)
            elif face_this_trial == f2_0v_5h:
                gaze_directions.append(str_0v_5h)
            elif face_this_trial == f2_0v_m5h:
                gaze_directions.append(str_0v_m5h)
            elif face_this_trial == f2_0v_10h:
                gaze_directions.append(str_0v_10h)
            elif face_this_trial == f2_0v_m10h:
                gaze_directions.append(str_0v_m10h)
            elif face_this_trial == f2_0v_15h:
                gaze_directions.append(str_0v_15h)
            elif face_this_trial == f2_0v_m15h:
                gaze_directions.append(str_0v_m15h)
            elif face_this_trial == f2_10v_0h:
                gaze_directions.append(str_10v_0h)
            elif face_this_trial == f2_m10v_0h:
                gaze_directions.append(str_m10v_0h)
            elif face_this_trial == f2_10v_5h:
                gaze_directions.append(str_10v_5h)
            elif face_this_trial == f2_10v_m5h:
                gaze_directions.append(str_10v_m5h)
            elif face_this_trial == f2_m10v_5h:
                gaze_directions.append(str_m10v_5h)
            elif face_this_trial == f2_m10v_m5h:
                gaze_directions.append(str_m10v_m5h)
            elif face_this_trial == f2_10v_10h:
                gaze_directions.append(str_10v_10h)
            elif face_this_trial == f2_10v_m10h:
                gaze_directions.append(str_10v_m10h)
            elif face_this_trial == f2_m10v_10h:
                gaze_directions.append(str_m10v_10h)
            elif face_this_trial == f2_m10v_m10h:
                gaze_directions.append(str_m10v_m10h)
            elif face_this_trial == f2_10v_15h:
                gaze_directions.append(str_10v_15h)
            elif face_this_trial == f2_10v_m15h:
                gaze_directions.append(str_10v_m15h)
            elif face_this_trial == f2_m10v_15h:
                gaze_directions.append(str_m10v_15h)
            elif face_this_trial == f2_m10v_m15h:
                gaze_directions.append(str_m10v_m15h)
                
            ### R-PEAK DETECTION ### 
            
            # For diastole trials we're going to be gathering 3 RR-intervals, so waiting for 4 R-peaks, then presenting
            # the stimulus just before the 5th R-peak. 
            
            # We'll do the same for systole trials to keep the ISIs similar, i.e., detecting 5 R-peaks, and presented 250ms after the 5th  
            
            for peak in range(5): 
                r_peak_time = listenforHR_basic()
                r_peak_times.append(r_peak_time)
                countdown.reset()
                r_peak_counter += 1 
                
                # A potential problem has been that when the threshold is crossed (i.e., an R-peak), the listenforHR_basic doesn't only send once, but during the
                # whole time that the signal crosses the threshold. So I'm introducing a delay during the sensing period, in order to pause everything
                # Until the signal has gone back below the threshold 
                
                if r_peak_counter < 5:
                    core.wait(0.5)
            
            while True: # waiting 250ms after R-peak has been detected
                if countdown.getTime() < (-0.250 + one_frame):
                    break
            
            win.flip()  
            countdown.reset()
            sendParallelTrigger(pport_address_spike, trigger_code_spike)
        
            while True:
                if countdown.getTime() < (-0.04 + one_frame): #75ms presentation of the stimulus
                    break
                
            win.flip() # remove stimulus
            
            ### FOR SYSTOLE TRIALS, CALCULATING THE MEAN RR-INTERVAL AND ADDING RR-INTERVALS TO LIST ### 
            
            for peak in range(len(r_peak_times) - 1): # going through the R-peak list
                rr_interval = r_peak_times[peak + 1] - r_peak_times[peak] # R-peak, index + 1, minus R-peak, index
                rr_intervals.append(rr_interval)
                
            mean_rr_int = mean(rr_intervals)
            
            
        elif ((rep_order == task_order_1) and (task_order_1[trial] == 2)) or ((rep_order == task_order_2) and (task_order_2[trial] == 2)): # a diastole trial 
        
            face_this_trial = random.choice(face_2_stim_d) # choosing which stimulus will be presented this trial
            face_2_stim_d.remove(face_this_trial) # removing the chosen simulus, so that it won't be presented this rep again 
            face_this_trial.draw()  # drawing the chosen stimulus 
            
            systole_diastole.append(2)
            
            if face_this_trial == f2_0v_0h:
                gaze_directions.append(str_0v_0h)
            elif face_this_trial == f2_0v_5h:
                gaze_directions.append(str_0v_5h)
            elif face_this_trial == f2_0v_m5h:
                gaze_directions.append(str_0v_m5h)
            elif face_this_trial == f2_0v_10h:
                gaze_directions.append(str_0v_10h)
            elif face_this_trial == f2_0v_m10h:
                gaze_directions.append(str_0v_m10h)
            elif face_this_trial == f2_0v_15h:
                gaze_directions.append(str_0v_15h)
            elif face_this_trial == f2_0v_m15h:
                gaze_directions.append(str_0v_m15h)
            elif face_this_trial == f2_10v_0h:
                gaze_directions.append(str_10v_0h)
            elif face_this_trial == f2_m10v_0h:
                gaze_directions.append(str_m10v_0h)
            elif face_this_trial == f2_10v_5h:
                gaze_directions.append(str_10v_5h)
            elif face_this_trial == f2_10v_m5h:
                gaze_directions.append(str_10v_m5h)
            elif face_this_trial == f2_m10v_5h:
                gaze_directions.append(str_m10v_5h)
            elif face_this_trial == f2_m10v_m5h:
                gaze_directions.append(str_m10v_m5h)
            elif face_this_trial == f2_10v_10h:
                gaze_directions.append(str_10v_10h)
            elif face_this_trial == f2_10v_m10h:
                gaze_directions.append(str_10v_m10h)
            elif face_this_trial == f2_m10v_10h:
                gaze_directions.append(str_m10v_10h)
            elif face_this_trial == f2_m10v_m10h:
                gaze_directions.append(str_m10v_m10h)
            elif face_this_trial == f2_10v_15h:
                gaze_directions.append(str_10v_15h)
            elif face_this_trial == f2_10v_m15h:
                gaze_directions.append(str_10v_m15h)
            elif face_this_trial == f2_m10v_15h:
                gaze_directions.append(str_m10v_15h)
            elif face_this_trial == f2_m10v_m15h:
                gaze_directions.append(str_m10v_m15h)
            
            ### R-PEAK DETECTION AND RR-INTERVAL CALCULATION ### 
            
            for peak in range(4): # gathering 4 R-peaks
                r_peak_time = listenforHR_basic()
                r_peak_times.append(r_peak_time)
                countdown.reset()
                r_peak_counter += 1
            
                
                if peak < r_peak_counter < 4: # again stopping multiple signals being sent per R-peak 
                    core.wait(0.5)

            ### CALCULATING RR-INTERVALS ### 
            
            for peak in range(len(r_peak_times) - 1): # going through the R-peak list
                rr_interval = r_peak_times[peak + 1] - r_peak_times[peak] # R-peak, index + 1, minus R-peak, index
                rr_intervals.append(rr_interval)
                
            mean_rr_int = mean(rr_intervals) # this is useful to know to check for faulty trials 
            
            predicted_rr_int = (rr_intervals[-1]*0.5) + (rr_intervals[-2]*0.3) + (rr_intervals[-3]*0.2)
            # ^ I'm calculating the predicted RR-interval by weighting the calculated RR-intervals, most recent with the highest weighting 
            
            ### DOING SOME BASIC ADJUSTMENTS BASED OFF THE MEAN RR-INTERVALS ### 
            
            if predicted_rr_int >= 1: # i.e., less than 60bpm
            
                while True:
                    if countdown.getTime() < (-predicted_rr_int + 0.150 + one_frame): # we present the stimulus 150ms before the next predicted R-peak
                        break
            
            elif (1 > predicted_rr_int) and (predicted_rr_int >= 0.86): #i.e., between 60 and 70bpm
                
                while True:
                    if countdown.getTime() < (-predicted_rr_int + 0.135 + one_frame): # present the stimulus 0.135ms before the next predicted R-peak
                        break
                    
            elif predicted_rr_int < 0.86: # i.e., greater than 70bpm
            
                while True: 
                    if countdown.getTime() < (-predicted_rr_int + 0.125 + one_frame): # present the stimulus 0.125ms before the next predicted R-peak 
                        break                
            
            win.flip() # Presenting the face 
            countdown.reset() # resetting the countdown timer now
            sendParallelTrigger(pport_address_spike, trigger_code_spike)
            
            while True:
                if countdown.getTime() < (-0.04 + one_frame):
                    break
                
            win.flip()
            
        ### GETTING GAZE DIRECTION RESPONSE ### 
        
        countdown.reset()
        
        circle = psychopy.visual.Circle(win, fillColor = "white", radius = 150)
        cross = psychopy.visual.shape.ShapeStim(win, fillColor = "red", vertices = 'cross', size = 25)
        
        # Function to check if the cross is within the circle
        def is_within_circle(cross_pos, circle_radius):
            return (cross_pos[0]**2 + cross_pos[1]**2) <= circle_radius**2
        
        cross_pos = [0,0]
        cross.setPos(cross_pos)
        
        while True:
            if countdown.getTime() < (-0.150 + one_frame): # A blank screen for 150ms after the presentation of the face 
                break
            
        cross.draw()
        circle.draw()
        win.flip()
        
     
        while True:
        
            if k.is_pressed("left"):
                new_pos = [cross_pos[0] - 3, cross_pos[1]]
            elif k.is_pressed("right"):
                new_pos = [cross_pos[0] + 3, cross_pos[1]]
            elif k.is_pressed("up"):
                new_pos = [cross_pos[0], cross_pos[1] + 3]
            elif k.is_pressed("down"):
                new_pos = [cross_pos[0], cross_pos[1] - 3]
            elif k.is_pressed("space"):
                break
            else:
                new_pos = cross_pos
            
            # Check if the new position is within the circle
            if is_within_circle(new_pos, 150):
                cross_pos = new_pos
                cross.setPos(cross_pos)
        
            # Draw the circle and the cross
            circle.draw()
            cross.draw()
            win.flip()
           
        response_coords_x = cross_pos[0]
        response_coords_y = cross_pos[1]
        
        distance_from_direct = (response_coords_x**2 + response_coords_y**2)**0.5 
        
        ### GETTING CONFIDENCE RESPONSE ### 
        
        countdown.reset()
      
        while True:
            if countdown.getTime() < (-0.25 + one_frame): # A blank screen for 250ms after the first perceptual rating 
                break
            
        # Followed by the confidence rating scale 
        
        triangle_position = 0
        space_pressed = False
        while space_pressed == False:
            
            # Draw the vertical line and confidence instructions
            vertical_line.draw()
            confidence_instructions.draw()
            high_confidence.draw()
            low_confidence.draw()
            triangle = visual.Polygon(win, edges = 3, radius = 15, pos = [-50, triangle_position])
            triangle.draw()

            
            for i in range(num_ticks):
                # Calculate the y position of each tick
                y_pos = -line_length / 2 + i * spacing
                # Create the horizontal tick
                tick = visual.Line(
                    win,
                    start=(-tick_length / 2, y_pos),
                    end=(tick_length / 2, y_pos),
                    lineColor='white')
                # Draw the tick
                tick.draw()
                
            win.flip()
            
            keys = event.waitKeys(keyList = ['up', 'down', 'space'])
            
            if (keys[0] == 'up') and (triangle_position != 125):
                triangle_position += 25
                
            elif (keys[0] == 'up') and (triangle_position == 125):
                triangle_position = 125
                
            elif (keys[0] == 'down') and (triangle_position != -125):
                triangle_position -= 25
            
            elif (keys[0] == 'down') and (triangle_position == -125):
                triangle_position == -25
                
            elif keys[0] == 'space':
                break 
            
        if triangle_position == 0:
            confidenceRating = 5
        elif triangle_position == 25:
            confidenceRating = 6
        elif triangle_position == 50:
            confidenceRating = 7
        elif triangle_position == 75:
            confidenceRating = 8
        elif triangle_position == 100:
            confidenceRating = 9
        elif triangle_position == 125:
            confidenceRating = 10
        elif triangle_position == -25:
            confidenceRating = 4
        elif triangle_position == -50:
            confidenceRating = 3
        elif triangle_position == -75:
            confidenceRating = 2
        elif triangle_position == -100:
            confidenceRating = 1
        elif triangle_position == -125:
            confidenceRating = 0 
                
        # Drawing the fixation cross for the next trial, will be shown for the next R-peak sensing period     
        
        win.flip()
        
        fixation_cross_horizontal.draw()
        fixation_cross_vertical.draw()
        
        # Recording the info for this trial 
        
        perceptual_responses_x.append(response_coords_x)
        perceptual_responses_y.append(response_coords_y)
        distances_from_direct.append(distance_from_direct)
        confidence_responses.append(confidenceRating)
        r_peak_times_lists.append(r_peak_times)
        mean_rr_intervals_list.append(mean_rr_int)
        
        if systole_diastole[-1] == 1: # a systole trial 
            predicted_rr_intervals_list.append(0) # just to fill the list (during systole trials we don't calculate a predicted rr int)
        
        elif systole_diastole[-1] == 2: # a diastole trial 
            predicted_rr_intervals_list.append(predicted_rr_int)
            
        win.flip()

### SAVING DATA FOR FACE 2 ### 

data_sheet = data_wb['face_2'] 

# Adding data to worksheet 

for trial in range(trials_per_block):
    # Adding gaze direction data for a trial 
    data_sheet.cell(trial+2, 1, gaze_directions[trial])
    # Adding systole/diastole data for a trial 
    data_sheet.cell(trial+2, 2, systole_diastole[trial])
    # Adding perceptual rating, x-coordinate
    data_sheet.cell(trial+2, 3, perceptual_responses_x[trial])
    # Adding perceptual rating, y-coordinate
    data_sheet.cell(trial+2, 4, perceptual_responses_y[trial])
    # Adding perceptual rating, distance from centre 
    data_sheet.cell(trial+2, 5, distances_from_direct[trial])
    # Adding confidence rating 
    data_sheet.cell(trial+2, 6, confidence_responses[trial])
    
    if systole_diastole[trial] == 1: # 5 r-peaks detected during a systole trial 
        
        data_sheet.cell(trial+2, 7, r_peak_times_lists[trial][0]) # first R-peak
        data_sheet.cell(trial+2, 8, r_peak_times_lists[trial][1]) # second R-peak...
        data_sheet.cell(trial+2, 9, r_peak_times_lists[trial][2]) 
        data_sheet.cell(trial+2, 10, r_peak_times_lists[trial][3]) 
        data_sheet.cell(trial+2, 11, r_peak_times_lists[trial][4]) 
    
    elif systole_diastole[trial] == 2: # 4 r-peaks detected during a diastole trial 
        
        data_sheet.cell(trial+2, 7, r_peak_times_lists[trial][0]) # first R-peak
        data_sheet.cell(trial+2, 8, r_peak_times_lists[trial][1]) # second R-peak...
        data_sheet.cell(trial+2, 9, r_peak_times_lists[trial][2]) 
        data_sheet.cell(trial+2, 10, r_peak_times_lists[trial][3]) 
        data_sheet.cell(trial+2, 13, predicted_rr_intervals_list[trial]) # predicted RR-interval added

    data_sheet.cell(trial+2, 12, mean_rr_intervals_list[trial]) # mean rr_interval added 
    
### EXPERIMENTAL BLOCK 3 ###

#Showing block complete dialogue

block_complete.draw()
win.flip()
event.waitKeys(keyList=['space'])
win.flip()

gaze_directions = [] # containing gaze directions per trial for a block

systole_diastole = [] # whether the trial is systole or diastole for a block
perceptual_responses_x = [] # perceptual responses per trial for a block, x-coordinate
perceptual_responses_y = [] # perceptual responses per trial for a block, y-coordinate 
distances_from_direct = [] # perceptual response, direct from centre 
confidence_responses = [] # confidence responses per trial for a block
r_peak_times_lists = [] # r-peak times per trial for a block
mean_rr_intervals_list = [] # rr-intervals per trial for a block 
predicted_rr_intervals_list = [] # Predicted RR-intervals per trial for a block 
 
countdown.reset()

fixation_cross_horizontal.draw()
fixation_cross_vertical.draw()

win.flip()

while True:
    if countdown.getTime() < (-2 + one_frame):
        break 

for rep in range(number_reps): # each gaze direction is shown five times per systole/diastole i.e., 5 reps per face     

    # Resetting the stimuli for this rep. Each rep consists of showing each gaze direction at systole and diastole
    
    face_3_stim = [f3_0v_0h, f3_0v_5h, f3_0v_m5h, f3_0v_10h, f3_0v_m10h, f3_0v_15h, f3_0v_m15h, f3_10v_0h, f3_m10v_0h, f3_10v_5h, f3_10v_m5h, f3_m10v_5h, f3_m10v_m5h, f3_10v_10h, f3_10v_m10h, f3_m10v_10h, f3_m10v_m10h, f3_10v_15h, f3_10v_m15h, f3_m10v_15h, f3_m10v_m15h]
    face_3_stim_d = [f3_0v_0h, f3_0v_5h, f3_0v_m5h, f3_0v_10h, f3_0v_m10h, f3_0v_15h, f3_0v_m15h, f3_10v_0h, f3_m10v_0h, f3_10v_5h, f3_10v_m5h, f3_m10v_5h, f3_m10v_m5h, f3_10v_10h, f3_10v_m10h, f3_m10v_10h, f3_m10v_m10h, f3_10v_15h, f3_10v_m15h, f3_m10v_15h, f3_m10v_m15h]
    face_3_stim_s = [f3_0v_0h, f3_0v_5h, f3_0v_m5h, f3_0v_10h, f3_0v_m10h, f3_0v_15h, f3_0v_m15h, f3_10v_0h, f3_m10v_0h, f3_10v_5h, f3_10v_m5h, f3_m10v_5h, f3_m10v_m5h, f3_10v_10h, f3_10v_m10h, f3_m10v_10h, f3_m10v_m10h, f3_10v_15h, f3_10v_m15h, f3_m10v_15h, f3_m10v_m15h]
    
    if order == 1: # the order in which this rep occurs 
        rep_order = order_1_structure[rep]
    elif order == 2:
        rep_order = order_2_structure[rep]
    
    # Entering the main stimulus presentation 
    for trial in range(rep_length):
        
        r_peak_counter = 0 # for tracking how many R-peaks have been detected this trial so far - remove when collecting actual data 
        
        r_peak_times = [] # the list that we'll use for each trial to measure R-peak times
        rr_intervals = [] 
        
        if ((rep_order == task_order_1) and (task_order_1[trial] == 1)) or ((rep_order == task_order_2) and task_order_2[trial] == 1): # a systole trial 
        
            face_this_trial = random.choice(face_3_stim_s) # choosing which stimulus will be presented this trial
            face_3_stim_s.remove(face_this_trial) # removing the chosen simulus, so that it won't be presented this rep again 
            face_this_trial.draw() # drawing the chosen stimulus 
            
            systole_diastole.append(1) # systole = 1, diastole = 2
            
            # Adding to the gaze direction list the face direction presented this trial
            
            if face_this_trial == f3_0v_0h:
                gaze_directions.append(str_0v_0h)
            elif face_this_trial == f3_0v_5h:
                gaze_directions.append(str_0v_5h)
            elif face_this_trial == f3_0v_m5h:
                gaze_directions.append(str_0v_m5h)
            elif face_this_trial == f3_0v_10h:
                gaze_directions.append(str_0v_10h)
            elif face_this_trial == f3_0v_m10h:
                gaze_directions.append(str_0v_m10h)
            elif face_this_trial == f3_0v_15h:
                gaze_directions.append(str_0v_15h)
            elif face_this_trial == f3_0v_m15h:
                gaze_directions.append(str_0v_m15h)
            elif face_this_trial == f3_10v_0h:
                gaze_directions.append(str_10v_0h)
            elif face_this_trial == f3_m10v_0h:
                gaze_directions.append(str_m10v_0h)
            elif face_this_trial == f3_10v_5h:
                gaze_directions.append(str_10v_5h)
            elif face_this_trial == f3_10v_m5h:
                gaze_directions.append(str_10v_m5h)
            elif face_this_trial == f3_m10v_5h:
                gaze_directions.append(str_m10v_5h)
            elif face_this_trial == f3_m10v_m5h:
                gaze_directions.append(str_m10v_m5h)
            elif face_this_trial == f3_10v_10h:
                gaze_directions.append(str_10v_10h)
            elif face_this_trial == f3_10v_m10h:
                gaze_directions.append(str_10v_m10h)
            elif face_this_trial == f3_m10v_10h:
                gaze_directions.append(str_m10v_10h)
            elif face_this_trial == f3_m10v_m10h:
                gaze_directions.append(str_m10v_m10h)
            elif face_this_trial == f3_10v_15h:
                gaze_directions.append(str_10v_15h)
            elif face_this_trial == f3_10v_m15h:
                gaze_directions.append(str_10v_m15h)
            elif face_this_trial == f3_m10v_15h:
                gaze_directions.append(str_m10v_15h)
            elif face_this_trial == f3_m10v_m15h:
                gaze_directions.append(str_m10v_m15h)

            
            ### R-PEAK DETECTION ### 
            
            # For diastole trials we're going to be gathering 3 RR-intervals, so waiting for 4 R-peaks, then presenting
            # the stimulus just before the 5th R-peak. 
            
            # We'll do the same for systole trials to keep the ISIs similar, i.e., detecting 5 R-peaks, and presented 250ms after the 5th  
            
            for peak in range(5): 
                r_peak_time = listenforHR_basic()
                r_peak_times.append(r_peak_time)
                countdown.reset()
                r_peak_counter += 1 
                
                # A potential problem has been that when the threshold is crossed (i.e., an R-peak), the listenforHR_basic doesn't only send once, but during the
                # whole time that the signal crosses the threshold. So I'm introducing a delay during the sensing period, in order to pause everything
                # Until the signal has gone back below the threshold 
                
                if r_peak_counter < 5:
                    core.wait(0.5)
            
            while True: # waiting 250ms after R-peak has been detected
                if countdown.getTime() < (-0.250 + one_frame):
                    break
            
            win.flip()  
            countdown.reset()
            sendParallelTrigger(pport_address_spike, trigger_code_spike)
        
            while True:
                if countdown.getTime() < (-0.04 + one_frame): #75ms presentation of the stimulus
                    break
                
            win.flip() # remove stimulus
            
            ### FOR SYSTOLE TRIALS, CALCULATING THE MEAN RR-INTERVAL AND ADDING RR-INTERVALS TO LIST ### 
            
            for peak in range(len(r_peak_times) - 1): # going through the R-peak list
                rr_interval = r_peak_times[peak + 1] - r_peak_times[peak] # R-peak, index + 1, minus R-peak, index
                rr_intervals.append(rr_interval)
                
            mean_rr_int = mean(rr_intervals)
            
            
        elif ((rep_order == task_order_1) and (task_order_1[trial] == 2)) or ((rep_order == task_order_2) and (task_order_2[trial] == 2)): # a diastole trial 
        
            face_this_trial = random.choice(face_3_stim_d) # choosing which stimulus will be presented this trial
            face_3_stim_d.remove(face_this_trial) # removing the chosen simulus, so that it won't be presented this rep again 
            face_this_trial.draw()  # drawing the chosen stimulus 
            
            systole_diastole.append(2)
            
            if face_this_trial == f3_0v_0h:
                gaze_directions.append(str_0v_0h)
            elif face_this_trial == f3_0v_5h:
                gaze_directions.append(str_0v_5h)
            elif face_this_trial == f3_0v_m5h:
                gaze_directions.append(str_0v_m5h)
            elif face_this_trial == f3_0v_10h:
                gaze_directions.append(str_0v_10h)
            elif face_this_trial == f3_0v_m10h:
                gaze_directions.append(str_0v_m10h)
            elif face_this_trial == f3_0v_15h:
                gaze_directions.append(str_0v_15h)
            elif face_this_trial == f3_0v_m15h:
                gaze_directions.append(str_0v_m15h)
            elif face_this_trial == f3_10v_0h:
                gaze_directions.append(str_10v_0h)
            elif face_this_trial == f3_m10v_0h:
                gaze_directions.append(str_m10v_0h)
            elif face_this_trial == f3_10v_5h:
                gaze_directions.append(str_10v_5h)
            elif face_this_trial == f3_10v_m5h:
                gaze_directions.append(str_10v_m5h)
            elif face_this_trial == f3_m10v_5h:
                gaze_directions.append(str_m10v_5h)
            elif face_this_trial == f3_m10v_m5h:
                gaze_directions.append(str_m10v_m5h)
            elif face_this_trial == f3_10v_10h:
                gaze_directions.append(str_10v_10h)
            elif face_this_trial == f3_10v_m10h:
                gaze_directions.append(str_10v_m10h)
            elif face_this_trial == f3_m10v_10h:
                gaze_directions.append(str_m10v_10h)
            elif face_this_trial == f3_m10v_m10h:
                gaze_directions.append(str_m10v_m10h)
            elif face_this_trial == f3_10v_15h:
                gaze_directions.append(str_10v_15h)
            elif face_this_trial == f3_10v_m15h:
                gaze_directions.append(str_10v_m15h)
            elif face_this_trial == f3_m10v_15h:
                gaze_directions.append(str_m10v_15h)
            elif face_this_trial == f3_m10v_m15h:
                gaze_directions.append(str_m10v_m15h)

            
            ### R-PEAK DETECTION AND RR-INTERVAL CALCULATION ### 
            
            for peak in range(4): # gathering 4 R-peaks
                r_peak_time = listenforHR_basic()
                r_peak_times.append(r_peak_time)
                countdown.reset()
                r_peak_counter += 1
            
                
                if peak < r_peak_counter < 4: # again stopping multiple signals being sent per R-peak 
                    core.wait(0.5)

            ### CALCULATING RR-INTERVALS ### 
            
            for peak in range(len(r_peak_times) - 1): # going through the R-peak list
                rr_interval = r_peak_times[peak + 1] - r_peak_times[peak] # R-peak, index + 1, minus R-peak, index
                rr_intervals.append(rr_interval)
                
            mean_rr_int = mean(rr_intervals) # this is useful to know to check for faulty trials 
            
            predicted_rr_int = (rr_intervals[-1]*0.5) + (rr_intervals[-2]*0.3) + (rr_intervals[-3]*0.2)
            # ^ I'm calculating the predicted RR-interval by weighting the calculated RR-intervals, most recent with the highest weighting 
            
            ### DOING SOME BASIC ADJUSTMENTS BASED OFF THE MEAN RR-INTERVALS ### 
            
            if predicted_rr_int >= 1: # i.e., less than 60bpm
            
                while True:
                    if countdown.getTime() < (-predicted_rr_int + 0.150 + one_frame): # we present the stimulus 150ms before the next predicted R-peak
                        break
            
            elif (1 > predicted_rr_int) and (predicted_rr_int >= 0.86): #i.e., between 60 and 70bpm
                
                while True:
                    if countdown.getTime() < (-predicted_rr_int + 0.135 + one_frame): # present the stimulus 0.135ms before the next predicted R-peak
                        break
                    
            elif predicted_rr_int < 0.86: # i.e., greater than 70bpm
            
                while True: 
                    if countdown.getTime() < (-predicted_rr_int + 0.125 + one_frame): # present the stimulus 0.125ms before the next predicted R-peak 
                        break                
            
            win.flip() # Presenting the face 
            countdown.reset() # resetting the countdown timer now
            sendParallelTrigger(pport_address_spike, trigger_code_spike)
            
            while True:
                if countdown.getTime() < (-0.04 + one_frame):
                    break
                
            win.flip()
            
        ### GETTING GAZE DIRECTION RESPONSE ### 
        
        countdown.reset()
        
        circle = psychopy.visual.Circle(win, fillColor = "white", radius = 150)
        cross = psychopy.visual.shape.ShapeStim(win, fillColor = "red", vertices = 'cross', size = 25)
        
        # Function to check if the cross is within the circle
        def is_within_circle(cross_pos, circle_radius):
            return (cross_pos[0]**2 + cross_pos[1]**2) <= circle_radius**2
        
        cross_pos = [0,0]
        cross.setPos(cross_pos)
        
        while True:
            if countdown.getTime() < (-0.150 + one_frame): # A blank screen for 150ms after the presentation of the face 
                break
            
        cross.draw()
        circle.draw()
        win.flip()
        
     
        while True:
        
            if k.is_pressed("left"):
                new_pos = [cross_pos[0] - 3, cross_pos[1]]
            elif k.is_pressed("right"):
                new_pos = [cross_pos[0] + 3, cross_pos[1]]
            elif k.is_pressed("up"):
                new_pos = [cross_pos[0], cross_pos[1] + 3]
            elif k.is_pressed("down"):
                new_pos = [cross_pos[0], cross_pos[1] - 3]
            elif k.is_pressed("space"):
                break
            else:
                new_pos = cross_pos
            
            # Check if the new position is within the circle
            if is_within_circle(new_pos, 150):
                cross_pos = new_pos
                cross.setPos(cross_pos)
        
            # Draw the circle and the cross
            circle.draw()
            cross.draw()
            win.flip()
           
        response_coords_x = cross_pos[0]
        response_coords_y = cross_pos[1]
        
        distance_from_direct = (response_coords_x**2 + response_coords_y**2)**0.5 
        
        ### GETTING CONFIDENCE RESPONSE ### 
        
        countdown.reset()
      
        while True:
            if countdown.getTime() < (-0.25 + one_frame): # A blank screen for 250ms after the first perceptual rating 
                break
            
        # Followed by the confidence rating scale 
        
        triangle_position = 0
        space_pressed = False
        while space_pressed == False:
            
            # Draw the vertical line and confidence instructions
            vertical_line.draw()
            confidence_instructions.draw()
            high_confidence.draw()
            low_confidence.draw()
            triangle = visual.Polygon(win, edges = 3, radius = 15, pos = [-50, triangle_position])
            triangle.draw()

            
            for i in range(num_ticks):
                # Calculate the y position of each tick
                y_pos = -line_length / 2 + i * spacing
                # Create the horizontal tick
                tick = visual.Line(
                    win,
                    start=(-tick_length / 2, y_pos),
                    end=(tick_length / 2, y_pos),
                    lineColor='white')
                # Draw the tick
                tick.draw()
                
            win.flip()
            
            keys = event.waitKeys(keyList = ['up', 'down', 'space'])
            
            if (keys[0] == 'up') and (triangle_position != 125):
                triangle_position += 25
                
            elif (keys[0] == 'up') and (triangle_position == 125):
                triangle_position = 125
                
            elif (keys[0] == 'down') and (triangle_position != -125):
                triangle_position -= 25
            
            elif (keys[0] == 'down') and (triangle_position == -125):
                triangle_position == -25
                
            elif keys[0] == 'space':
                break 
            
        if triangle_position == 0:
            confidenceRating = 5
        elif triangle_position == 25:
            confidenceRating = 6
        elif triangle_position == 50:
            confidenceRating = 7
        elif triangle_position == 75:
            confidenceRating = 8
        elif triangle_position == 100:
            confidenceRating = 9
        elif triangle_position == 125:
            confidenceRating = 10
        elif triangle_position == -25:
            confidenceRating = 4
        elif triangle_position == -50:
            confidenceRating = 3
        elif triangle_position == -75:
            confidenceRating = 2
        elif triangle_position == -100:
            confidenceRating = 1
        elif triangle_position == -125:
            confidenceRating = 0 
                
        # Drawing the fixation cross for the next trial, will be shown for the next R-peak sensing period     
        
        win.flip()
        
        fixation_cross_horizontal.draw()
        fixation_cross_vertical.draw()
        
        # Recording the info for this trial 
        
        perceptual_responses_x.append(response_coords_x)
        perceptual_responses_y.append(response_coords_y)
        distances_from_direct.append(distance_from_direct)
        confidence_responses.append(confidenceRating)
        r_peak_times_lists.append(r_peak_times)
        mean_rr_intervals_list.append(mean_rr_int)
        
        if systole_diastole[-1] == 1: # a systole trial 
            predicted_rr_intervals_list.append(0) # just to fill the list (during systole trials we don't calculate a predicted rr int)
        
        elif systole_diastole[-1] == 2: # a diastole trial 
            predicted_rr_intervals_list.append(predicted_rr_int)
            
        win.flip()

### SAVING DATA FOR FACE 3 ### 

data_sheet = data_wb['face_3'] 

# Adding data to worksheet 

for trial in range(trials_per_block):
    # Adding gaze direction data for a trial 
    data_sheet.cell(trial+2, 1, gaze_directions[trial])
    # Adding systole/diastole data for a trial 
    data_sheet.cell(trial+2, 2, systole_diastole[trial])
    # Adding perceptual rating, x-coordinate
    data_sheet.cell(trial+2, 3, perceptual_responses_x[trial])
    # Adding perceptual rating, y-coordinate
    data_sheet.cell(trial+2, 4, perceptual_responses_y[trial])
    # Adding perceptual rating, distance from centre 
    data_sheet.cell(trial+2, 5, distances_from_direct[trial])
    # Adding confidence rating 
    data_sheet.cell(trial+2, 6, confidence_responses[trial])
    
    if systole_diastole[trial] == 1: # 5 r-peaks detected during a systole trial 
        
        data_sheet.cell(trial+2, 7, r_peak_times_lists[trial][0]) # first R-peak
        data_sheet.cell(trial+2, 8, r_peak_times_lists[trial][1]) # second R-peak...
        data_sheet.cell(trial+2, 9, r_peak_times_lists[trial][2]) 
        data_sheet.cell(trial+2, 10, r_peak_times_lists[trial][3]) 
        data_sheet.cell(trial+2, 11, r_peak_times_lists[trial][4]) 
    
    elif systole_diastole[trial] == 2: # 4 r-peaks detected during a diastole trial 
        
        data_sheet.cell(trial+2, 7, r_peak_times_lists[trial][0]) # first R-peak
        data_sheet.cell(trial+2, 8, r_peak_times_lists[trial][1]) # second R-peak...
        data_sheet.cell(trial+2, 9, r_peak_times_lists[trial][2]) 
        data_sheet.cell(trial+2, 10, r_peak_times_lists[trial][3]) 
        data_sheet.cell(trial+2, 13, predicted_rr_intervals_list[trial]) # predicted RR-interval added

    data_sheet.cell(trial+2, 12, mean_rr_intervals_list[trial]) # mean rr_interval added 
    
data_wb.save(save_file)
data_wb.close()

win.close()
core.quit()
        