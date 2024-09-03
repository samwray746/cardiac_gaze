import openpyxl
import os.path as op

def save_file_modification(subject_id):
    data_path = 'D:\\PhD\\cardiac_gaze\\data'
    subject_id_file_name = subject_id + 'gabor_task_data.xlsx'
    save_file = op.join(data_path, subject_id_file_name)

    data_wb = openpyxl.Workbook() # opening the excel file now 

    sheet_names = ['training', 'block_1', 'block_2', 'block_3', 'summary_data']

    heading_strings = ['trial_number', 'trial_type', 'angle_offset', 'left_right', 'clockwise_anticlockwise', 'correct_incorrect', 'r_peak_time_1', 'r_peak_time_2', 'r_peak_time_3', 'r_peak_time_4', 'r_peak_time_5', 'mean_rr_int', 'predicted_rr_int']
    
    for ws_name in range(len(sheet_names)):
        data_wb.create_sheet(sheet_names[ws_name]) # creating a sheet to save each block 
        
        data_sheet = data_wb[sheet_names[ws_name]]
        
        for header in range(len(heading_strings)): # adding the headers
            data_sheet.cell(1, header+1, heading_strings[header])
    
    return data_wb, save_file
