# -*- coding: utf-8 -*-
"""
Created on Thu Jul 11 14:10:30 2024

@author: bsms9zh8
"""
import openpyxl
import os.path as op

def save_file_modification(subject_id):
    data_path = 'D:\\PhD\\cardiac_gaze\\data'
    subject_id_file_name = subject_id + '_face_task_data.xlsx'
    save_file = op.join(data_path, subject_id_file_name)

    data_wb = openpyxl.Workbook() # opening the excel file now 

    faces_strings = ['face_1', 'face_2', 'face_3']
    variable_strings = ['gaze_angle', 'systole_diastole', 'perceptual_response_x', 'perceptual_response_y', 'distance_from_centre', 'confidence_response', 'r_peak_time_1', 'r_peak_time_2', 'r_peak_time_3', 'r_peak_time_4', 'r_peak_time_5', 'mean_rr_int', 'predicted_rr_int']

    for ws_name in range(len(faces_strings)):
        data_wb.create_sheet(faces_strings[ws_name]) # creating a sheet to save each block 
        
        data_sheet = data_wb[faces_strings[ws_name]]
        
        for header in range(len(variable_strings)): # adding the headers
            data_sheet.cell(1, header+1, variable_strings[header])
    
    return data_wb, save_file
